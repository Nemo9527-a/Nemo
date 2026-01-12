# -*- coding: utf-8 -*-
# @Time    : 2025/3/3  16:05
# @Author  : Nemo
# @FileName: excel_util.py
# @Software: PyCharm
"""
    Description:EXCEL读写封装
"""
import time
from LEDsetting.config.conf import cm
from openpyxl import load_workbook
from LEDsetting.common.log_util import log


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.sheet_name = sheet_name  # 表单名
        self.path = cm.DATA_FILE + "/" + file_name # 获取excel文件路径
        self.workbook = None
        self.sheet = None

    @classmethod
    def load_workbook(cls, file_path):
        """
        打开excel文件
        :param file_path:
        :return:
        """
        return load_workbook(file_path)

    def open_excel(self):
        """
        打开sheet页
        :return:
        """
        self.workbook = self.load_workbook(self.path)
        self.sheet = self.workbook[self.sheet_name]

    def close_excel(self):
        """
        关闭excel
        :return:
        """
        if self.workbook:
            self.workbook.close()

    def get_sheet_name(self):
        """
        获取所有sheet名
        :return:
        """
        return self.workbook.sheetnames

    def read(self):
        """
        读取EXCEL数据
        :return:
        """
        self.open_excel()
        # 获取最大行和最大列
        row_max = self.sheet.max_row
        col_max = self.sheet.max_column
        cases = list()
        for i in range(2, row_max + 1):
            case = dict()
            case['port_name'] = self.sheet_name  # 读取port_name
            case['case_title'] = self.sheet.cell(i, 1).value  # 读取case_id
            case['protocol'] = self.sheet.cell(i, 2).value  # 读取protocol
            case['url'] = self.sheet.cell(i, 3).value  # 读取url
            case['method'] = self.sheet.cell(i, 4).value  # 读取method
            case['header'] = self.sheet.cell(i, 5).value  # 读取header
            case['data'] = self.sheet.cell(i, 6).value  # 读取data
            case['excepted'] = self.sheet.cell(i, 7).value  # 读取excepted
            cases.append(case)
        self.close_excel()
        return cases

    def write(self, row, col, value):
        """
        将结果回写至EXCEL
        :param row:
        :param col:
        :param value:
        :return:
        """
        self.open_excel()
        self.sheet.cell(row, col).value = value  # 写入值到单元格里面去
        self.workbook.save(self.path)  # 要记得保存，同时要记得Excel要关闭状态
        self.close_excel()

    def write_result_to_excel(self, case_title, result_test):
        """
        将测试结果写入 Excel 的 result 列
        :param case_title: 用例标题，用于定位行
        :param result_test: 测试结果
        """
        # 加载 Excel 文件
        self.open_excel()
        # 遍历 Excel 文件，找到对应的用例行
        for row in range(2, self.sheet.max_row + 1):  # 从第 2 行开始遍历
            if self.sheet.cell(row=row, column=1).value == case_title:  # 第 1 列是 case_title
                # 将结果写入 result 列
                self.sheet.cell(row=row, column=8).value = result_test
                time_now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
                self.sheet.cell(row=row, column=9).value = time_now
                log.info(f"第{row}行执行结果成功写入excel")
                break
            else:
                log.info(f"第{row}行未找到匹配行!")

        # 保存 Excel 文件
        self.workbook.save(self.path)
        self.close_excel()



if __name__ == '__main__':
    D = DoExcel('param.xlsx', '查询所有发送器').read()
    print("读取的excel数据：", D)

    # DoExcel('param.xlsx', '查询已初始化客户端').write_result_to_excel('command参数为空', 11111)
    # url = 'https://www.xxx.com/login'
    # for i in D:
    #     data = eval(i['data'])
    #     case_id = i['case_id'] + 1
    #     print(case_id)
    #     r = requests.request("post", url, data=data)
    #     DoExcel('Login_Data.xlsx', 'login').write_excel(case_id, 6, r.text)
    #     print("状态码:{},\n返回信息:{}，\n相应内容编码:{}".format(r.status_code, r.text, r.encoding))
